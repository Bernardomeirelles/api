# requirements.txt:
# streamlit
# pandas
# requests
# openpyxl

import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import json

st.set_page_config(page_title="Deskfy API - Exportação", layout="centered")
st.title("Deskfy API - Consulta e Exportação")

with st.sidebar:
    st.header("Credenciais e Configuração")
    st.session_state.api_key = "47200d78-1f10-44ba-a8e4-187a7d35e3bd"
    st.markdown(":white_check_mark: API Key carregada com sucesso.")

def formatar_excel(writer, sheet_name):
    ws = writer.sheets[sheet_name]
    for col_num, column_cells in enumerate(ws.iter_cols(1, ws.max_column), 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 70)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

def dict_flatten(d, prefix=''):
    out = {}
    for k, v in d.items():
        col = f"{prefix}{k}"
        if isinstance(v, dict):
            for subk, subv in dict_flatten(v, f"{col}.").items():
                out[subk] = subv
        elif isinstance(v, list):
            # Lista de dicts: transforma em JSON string amigável
            if all(isinstance(i, dict) for i in v):
                out[col] = "; ".join(json.dumps(i, ensure_ascii=False) for i in v)
            else:
                out[col] = ", ".join(str(i) for i in v)
        elif v is None:
            out[col] = ""
        else:
            out[col] = v
    return out

opcao = st.radio("Selecione a operação:", ["Consultar Relatórios", "Consultar Detalhes da Solicitação"])

# ---- OPÇÃO 1: RELATÓRIOS ----
if opcao == "Consultar Relatórios":
    st.subheader("Parâmetros da Consulta")
    col1, col2 = st.columns(2)
    with col1:
        initial_date = st.date_input("Data Inicial")
    with col2:
        end_date = st.date_input("Data Final")
    briefing_id = st.text_input("Briefing ID")
    board_name = st.text_input("Board (opcional)")
    column_name = st.text_input("Coluna (opcional)")

    if st.button("Consultar Relatórios"):
        with st.spinner("Consultando API da Deskfy..."):
            if not st.session_state.api_key or not initial_date or not end_date or not briefing_id:
                st.warning("Preencha todos os campos obrigatórios.")
            else:
                url = "https://service-api.deskfy.io/v1/reports/workflow"
                params = {
                    "initialDate": initial_date.strftime("%Y-%m-%d"),
                    "endDate": end_date.strftime("%Y-%m-%d"),
                    "briefingId": briefing_id
                }
                if board_name:
                    params["boardName"] = board_name
                if column_name:
                    params["columnName"] = column_name

                response = requests.get(url, headers={"x-api-key": st.session_state.api_key}, params=params)

                if response.status_code == 200:
                    data = response.json()
                    if not data:
                        st.info("Nenhum dado encontrado.")
                    else:
                        # Garante que cada item do relatório seja "flattened"
                        data_flat = []
                        for item in data:
                            data_flat.append(dict_flatten(item))
                        df = pd.DataFrame(data_flat)
                        st.success("Dados obtidos com sucesso.")
                        st.dataframe(df)

                        output_excel = BytesIO()
                        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name="Relatório")
                            formatar_excel(writer, "Relatório")
                        output_excel.seek(0)

                        st.download_button("Baixar Excel", data=output_excel, file_name="relatorio_deskfy.xlsx")
                        output_csv = df.to_csv(index=False).encode("utf-8")
                        st.download_button("Baixar CSV", data=output_csv, file_name="relatorio_deskfy.csv")
                else:
                    st.error(f"Erro {response.status_code}: {response.text}")

# ---- OPÇÃO 2: DETALHES DA SOLICITAÇÃO ----
if opcao == "Consultar Detalhes da Solicitação":
    st.subheader("Detalhes da Solicitação")
    task_id = st.text_input("Task ID da Solicitação")

    if st.button("Consultar Detalhes"):
        with st.spinner("Buscando dados da solicitação..."):
            if not (st.session_state.api_key and task_id):
                st.warning("Preencha a API Key e o Task ID.")
            else:
                url = "https://service-api.deskfy.io/v1/reports/workflow/task-details"
                response = requests.get(url, headers={"x-api-key": st.session_state.api_key}, params={"taskId": task_id})

                if response.status_code == 200:
                    data = response.json()
                    st.success("Dados carregados com sucesso.")
                    st.json(data)

                    linha = {}
                    # Pega tudo do nível raiz e campos aninhados
                    for key, value in data.items():
                        if isinstance(value, dict):
                            for subk, subv in dict_flatten(value, prefix=f"{key}.").items():
                                linha[subk] = subv
                        elif isinstance(value, list):
                            if all(isinstance(i, dict) for i in value):
                                linha[key] = "; ".join(json.dumps(i, ensure_ascii=False) for i in value)
                            else:
                                linha[key] = ", ".join(str(i) for i in value)
                        elif value is None:
                            linha[key] = ""
                        else:
                            linha[key] = value

                    # Link Deskfy amigável (se existir id)
                    if "solicitacao.id" in linha:
                        linha["link_deskfy"] = f"https://app.deskfy.io/task/{linha['solicitacao.id']}"

                    df = pd.DataFrame([linha])
                    st.dataframe(df)

                    output = BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Detalhes")
                        formatar_excel(writer, "Detalhes")
                    output.seek(0)

                    st.download_button("Baixar Excel da Solicitação", data=output, file_name=f"detalhes_{linha.get('solicitacao.codigo', task_id)}.xlsx")
                else:
                    st.error(f"Erro Deskfy {response.status_code}: {response.text}")
