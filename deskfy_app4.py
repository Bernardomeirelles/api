

## -------------------------------------------------------------------------------------------------------


# requirements.txt:
# streamlit
# pandas
# requests
# openpyxl

import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import json
import hashlib

st.set_page_config(page_title="Deskfy API - Exportação", layout="centered")
st.title("Deskfy API - Consulta e Exportação")

# --- Campos esperados ---
CAMPOS_BEBIDAS = [
    "chopp_(digitar_nome,_mililitros_e_preco)", "beck´s_600ml", "corona_600ml", "stella_artois_600ml", "spaten_600ml",
    "original_600ml", "serramalte_600ml", "budweiser_600ml", "brahma_duplo_malte_600ml", "brahma_chopp_600ml",
    "skol_600ml", "bohemia_600ml", "antarctica_600ml", "heineken_600ml", "eisenbahn_600ml", "amstel_600ml",
    "petra_600ml", "itaipava_600ml", "devassa_600ml", "imperio_600ml", "patagonia_ipa_600ml",
    "patagonia_amber_lager_600ml", "colorado_indica_600ml", "colorado_appia_600ml", "colorado_cauim_600ml",
    "colorado_ribeirao_lager_600ml", "corona_ln", "beck´s_ln", "stella_pure_gold_ln", "stella_artois_ln",
    "spaten_ln", "budweiser_ln", "malzbier_ln", "heineken_ln", "corona_cero_ln", "heineken_zero_ln",
    "budweiser_zero_ln", "patagonia_ln", "colorado_ribeirao_ln", "original_1l", "budweiser_1l",
    "brahma_duplo_malte_1l", "brahma_chopp_1l", "bohemia_1l", "skol_1l", "antarctica_1l", "itaipava_1l",
    "amstel_1l", "skol_300ml", "brahma_chopp_300ml", "bohemia_300ml", "antarctica_pilsen_300ml",
    "brahma_duplo_malte_300ml", "original_300ml", "budweiser_300ml", "amstel_300ml", "devassa_300ml",
    "petra_300ml", "itaipava_300ml", "beats_ln", "beats_lata", "mike's_ln", "mike's_lata",
    "guarana_antarctica_lata", "guarana_antarctica_zero_lata", "pepsi_lata", "pepsi_black_lata",
    "tonica_antarctica_lata", "coca_cola_lata", "red_bull", "gatorade_500ml", "h20h_500ml",
    "guarana_antarctica_1l", "guarana_antarctica_zero_1l", "pepsi_1l", "pepsi_black_1l",
    "coca_cola_1l", "coca_cola_zero_1l","o_pdv_deseja_fazer_alteracoes_nas_secoes_de_comida?_se_sim,_digite_aqui.",
    "os_itens_novos_do_cardapio_tem_codigo?_se_sim,_digite_aqui.","outros_produtos_nao_listados_(digitar_marca_e_preco)"
]

ORDEM_FORMULARIO = [
    "briefing.unb", "briefing.sala", "briefing.setor", "briefing.codigo_pdv", "briefing.nome_pdv",
    "briefing.o_pdv_ja_tem_cardapio?",
    "briefing.o_que_foi_negociado_com_o_pdv?_e_possivel_selecionar_os_dois.",
    "briefing.trabalho_a_ser_feito",
    "briefing.tipo_de_cardapio",
    "briefing.ja_produzimos_esse_cardapio_antes?",
    "briefing.vip/premium",
    "briefing.material",
    "briefing.folha_(no_minimo_2_folhas_para_encadernar;_nao_e_possivel_encadernar_placa_ps)",
    "briefing.medidas_(largura_x_altura)",
    "briefing.copias",
    "briefing.design_ou_impressao_(casos_de_apenas_impressao_tambem_precisam_estar_de_acordo_com_nossos_guidelines)",
    "briefing.marca_foco",
    "briefing.embalagens_na_capa_(selecione_quantas_quiser)",
    *["briefing." + campo for campo in CAMPOS_BEBIDAS]
]

# --- Sidebar com CSS e Navegação ---
with st.sidebar:
    st.markdown("""
    <style>
    [data-testid="stSidebar"] {
        min-width: 350px;
        width: 350px;
    }
    [data-testid="stSidebar"] .css-1wvake5 {
        white-space: normal;
    }
    div[data-baseweb="select"] > div {
        background-color: black !important;
        color: white !important;
    }
    ul[role="listbox"] > li {
        background-color: white !important;
        color: black !important;
    }
    ul[role="listbox"] > li[aria-selected="true"] {
        background-color: black !important;
        color: white !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.header("Credenciais e Configuração")
    st.session_state.api_key = "47200d78-1f10-44ba-a8e4-187a7d35e3bd"
    st.success("✅ API Key carregada com sucesso.")
    st.text_input("🔐 API Key (visualização)", value=st.session_state.api_key, disabled=True)
    st.caption("A chave acima está em uso e não pode ser editada.")
    pagina = st.selectbox("📄 Escolha a funcionalidade:", ["Cardápios Solicitados", "Outros Relatórios"])

# --- Sessão para salvar consultas detalhadas ---
if "consultas_salvas" not in st.session_state:
    st.session_state.consultas_salvas = {}

# --- Funções auxiliares ---
def dict_flatten(d, prefix=''):
    out = {}
    for k, v in d.items():
        col = f"{prefix}{k}"
        if isinstance(v, dict):
            out.update(dict_flatten(v, prefix=f"{col}."))
        elif isinstance(v, list):
            out[col] = "; ".join(json.dumps(i, ensure_ascii=False) if isinstance(i, dict) else str(i) for i in v)
        else:
            out[col] = v if v is not None else ""
    return out

def formatar_excel(writer, sheet_name):
    ws = writer.sheets[sheet_name]
    for col_num, column_cells in enumerate(ws.iter_cols(1, ws.max_column), 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 70)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

def exibir_tabela_formatada(linha):
    st.markdown("### 📋 Tabela Formatada (Cópia)")
    table_md = "| Campo | Valor |\n|-------|-------|\n"
    for campo in ORDEM_FORMULARIO:
        valor = linha.get(campo, "Não Alterado" if campo.startswith("briefing.") else "")
        campo_legivel = campo.replace("briefing.", "").replace("_", " ").capitalize()
        table_md += f"| {campo_legivel} | {valor if valor else 'Não Alterado'} |\n"
    st.code(table_md, language="markdown")

def hash_dict(d):
    return hashlib.md5(json.dumps(d, sort_keys=True).encode()).hexdigest()

def salvar_consulta(task_id, linha):
    hash_atual = hash_dict(linha)
    if task_id not in st.session_state.consultas_salvas or st.session_state.consultas_salvas[task_id]['hash'] != hash_atual:
        st.session_state.consultas_salvas[task_id] = {"hash": hash_atual, "dados": linha}

# === Página: Cardápios Solicitados ===
if pagina == "Cardápios Solicitados":
    st.subheader("Detalhes da Solicitação")
    task_id = st.text_input("Task ID da Solicitação")

    if st.button("Consultar Detalhes"):
        with st.spinner("Buscando dados da solicitação..."):
            if not (st.session_state.api_key and task_id):
                st.warning("Preencha a API Key e o Task ID.")
            else:
                url = "https://service-api.deskfy.io/v1/reports/workflow/task-details"
                response = requests.get(url, headers={"x-api-key": st.session_state.api_key}, params={"taskId": task_id})

                if response.status_code == 200:
                    data = response.json()
                    st.success("Dados carregados com sucesso.")

                    linha = dict_flatten(data)
                    if "solicitacao.id" in linha:
                        linha["link_deskfy"] = f"https://app.deskfy.io/task/{linha['solicitacao.id']}"

                    salvar_consulta(task_id, linha)

                    # Coleta os valores conforme a ordem definida
                    valores_ordenados = [str(linha.get(campo, "")) for campo in ORDEM_FORMULARIO]

                    # Converte os nomes dos campos para cabeçalhos legíveis
                    cabecalhos_legiveis = [campo.replace("briefing.", "").replace("_", " ").capitalize() for campo in ORDEM_FORMULARIO]

                    # Cria DataFrame apenas com os campos da ORDEM_FORMULARIO (incluindo bebidas)
                    df_exibicao = pd.DataFrame([valores_ordenados], columns=cabecalhos_legiveis)

                    # Mostra na interface do app
                    st.dataframe(df_exibicao)

                    # Exporta para Excel com apenas os campos esperados
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df_exibicao.to_excel(writer, index=False, sheet_name="Detalhes")
                        formatar_excel(writer, "Detalhes")
                    output.seek(0)

                    # Botão para download
                    st.download_button(
                        "📥 Baixar Excel da Solicitação",
                        data=output,
                        file_name=f"detalhes_{linha.get('solicitacao.codigo', task_id)}.xlsx"
                    )


                    exibir_tabela_formatada(linha)

                    df_export = pd.DataFrame([valores_ordenados], columns=cabecalhos_legiveis)
                   
                else:
                    st.error(f"Erro Deskfy {response.status_code}: {response.text}")

    if st.button("Ver Histórico de Consultas"):
        st.markdown(f"### Total de Consultas: {len(st.session_state.consultas_salvas)}")
        for tid, entry in st.session_state.consultas_salvas.items():
            with st.expander(f"🔎 {tid}"):
                exibir_tabela_formatada(entry["dados"])
                st.code(json.dumps(entry["dados"], indent=2, ensure_ascii=False), language="json")

# === Página: Outros Relatórios ===
if pagina == "Outros Relatórios":
    st.header("📊 Consultar Relatórios")
    col1, col2 = st.columns(2)
    with col1:
        initial_date = st.date_input("Data Inicial")
    with col2:
        end_date = st.date_input("Data Final")
    briefing_id = st.text_input("Briefing ID")
    board_name = st.text_input("Board (opcional)")
    column_name = st.text_input("Coluna (opcional)")

    if st.button("Consultar Relatórios"):
        with st.spinner("Consultando API da Deskfy..."):
            if not st.session_state.api_key or not initial_date or not end_date or not briefing_id:
                st.warning("Preencha todos os campos obrigatórios.")
            else:
                url = "https://service-api.deskfy.io/v1/reports/workflow"
                params = {
                    "initialDate": initial_date.strftime("%Y-%m-%d"),
                    "endDate": end_date.strftime("%Y-%m-%d"),
                    "briefingId": briefing_id
                }
                if board_name:
                    params["boardName"] = board_name
                if column_name:
                    params["columnName"] = column_name

                response = requests.get(url, headers={"x-api-key": st.session_state.api_key}, params=params)

                if response.status_code == 200:
                    data = response.json()
                    if not data:
                        st.info("Nenhum dado encontrado.")
                    else:
                        data_flat = [dict_flatten(item) for item in data]
                        df = pd.DataFrame(data_flat)
                        st.success("Dados obtidos com sucesso.")
                        st.dataframe(df)

                        output_excel = BytesIO()
                        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name="Relatório")
                            formatar_excel(writer, "Relatório")
                        output_excel.seek(0)

                        st.download_button("📥 Baixar Excel", data=output_excel, file_name="relatorio_deskfy.xlsx")
                        st.download_button("📄 Baixar CSV", data=df.to_csv(index=False).encode("utf-8"), file_name="relatorio_deskfy.csv")
                else:
                    st.error(f"Erro {response.status_code}: {response.text}")
