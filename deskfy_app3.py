import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

st.set_page_config(page_title="Deskfy API - Consulta e Exportação", layout="centered")
st.title("Deskfy API - Consulta e Exportação")

COLUNAS_MODELO = [
    "COLUNA DESKFY", "STATUS", "DESCRIÇÃO STATUS", "GEO", "UNB", "SALA", "SETOR", "COMERCIAL", "OPERAÇÃO", "SOLICITANTE",
    "PDV", "NOME FANTASIA", "LINK DESKFY", "DATA SOLICITAÇÃO", "MÊS SOLICITAÇÃO", "ANO SOLICITAÇÃO", "PDV TEM CARDÁPIO",
    "DESIGN VS IMPRESSÃO", "MOTIVO SOLICITAÇÃO", "DEMANDA", "TIPO CARDÁPIO", "JÁ PRODUZIMOS ANTES", "VIP/PREMIUM",
    "MATERIAL", "FOLHA", "CÓPIAS", "MARCA FOCO", "DATA APROVAÇÃO SOLICITAÇÃO", "DATA INÍCIO PRODUÇÃO", "DATA 1A ENTREGA",
    "DATA REVISÃO ARTE", "DATA APROVAÇÃO ARTE", "DATA ENVIO IMPRESSÃO", "AJUSTES", "PÁGINAS", "DATA ENTREGUE",
    "MÊS ENTREGUE", "ANO ENTREGUE", "TEMPO APROVAÇÃO SOLICITAÇÃO", "TEMPO TEMPO FILA ARTE", "TEMPO PRODUÇÃO ARTE",
    "TEMPO REVISÃO ARTE", "TEMPO ATÉ APROVAÇÃO ARTE FINAL", "TEMPO FILA IMPRESSÃO", "TEMPO ENTREGA CDD", "SLA TOTAL",
    "IDADE MÉDIA", "DATA CANCELAMENTO", "MOTIVO CANCELAMENTO", "EMPRESA", "DESIGNER"
]

CAMPOS_EXCEL_FORMULA = [
    "STATUS", "DESCRIÇÃO STATUS", "COMERCIAL", "OPERAÇÃO", "SOLICITANTE",
    "MÊS SOLICITAÇÃO", "ANO SOLICITAÇÃO", "MÊS ENTREGUE", "ANO ENTREGUE",
    "TEMPO APROVAÇÃO SOLICITAÇÃO", "TEMPO TEMPO FILA ARTE", "TEMPO PRODUÇÃO ARTE",
    "TEMPO REVISÃO ARTE", "TEMPO ATÉ APROVAÇÃO ARTE FINAL", "TEMPO FILA IMPRESSÃO",
    "TEMPO ENTREGA CDD", "SLA TOTAL", "IDADE MÉDIA"
]

# Mapeamentos (mantidos como no seu código original)
mapeamento_briefing = {
    "PDV TEM CARDÁPIO": ["o_pdv_ja_tem_cardapio?", "pdv_tem_cardapio", "pdv ja tem cardapio"],
    "DESIGN VS IMPRESSÃO": ["design_ou_impressao_(casos_de_apenas_impressao_tambem_precisam_estar_de_acordo_com_nossos_guidelines)",
                            "design_ou_impressao", "design vs impressão"],
    "MOTIVO SOLICITAÇÃO": ["o_que_foi_negociado_com_o_pdv?_e_possivel_selecionar_os_dois.",
                           "motivo_solicitacao", "motivo da solicitação"],
    "DEMANDA": ["trabalho_a_ser_feito", "demanda", "tipo de demanda"],
    "TIPO CARDÁPIO": ["tipo_de_cardapio", "tipo_cardapio", "tipo de cardapio"],
    "JÁ PRODUZIMOS ANTES": ["ja_produzimos_esse_cardapio_antes?", "ja_produzimos_antes", "já produzimos esse cardápio antes?"],
    "VIP/PREMIUM": ["vip_premium", "vip/premium"],
    "MATERIAL": ["folha_(no_minimo_2_folhas_para_encadernar;_nao_e_possivel_encadernar_placa_ps)", "material"],
    "FOLHA": ["medidas_(largura_x_altura)", "folha", "formato"],
    "CÓPIAS": ["copias", "cópias"],
    "MARCA FOCO": ["marca_foco", "marca foco"],
    "EMPRESA": ["empresa"],
    "DESIGNER": ["designer"],
    "SALA": ["sala"],
    "SETOR": ["setor"],
    "PÁGINAS": ["paginas", "número de páginas"]
}

mapeamento_sol = {
    "COLUNA DESKFY": ["codigo", "coluna deskfy"],
    "STATUS": ["status"],
    "DESCRIÇÃO STATUS": ["colunaatual", "descrição status"],
    "OPERAÇÃO": ["board", "operação"],
    "SOLICITANTE": ["solicitante"],
    "PDV": ["codigo_pdv", "pdv"],
    "NOME FANTASIA": ["nome_pdv", "nome fantasia"],
    "LINK DESKFY": ["id"],  # será formatado depois
    "DATA SOLICITAÇÃO": ["dt_cadastro"],
    "AJUSTES": ["ajustes"],
    "DATA APROVAÇÃO SOLICITAÇÃO": ["dt_aprovacao_solicitacao"],
    "DATA INÍCIO PRODUÇÃO": ["dt_inicio_producao"],
    "DATA 1A ENTREGA": ["dt_1a_entrega"],
    "DATA REVISÃO ARTE": ["dt_revisao_arte"],
    "DATA APROVAÇÃO ARTE": ["dt_aprovacao_arte"],
    "DATA ENVIO IMPRESSÃO": ["dt_envio_impressao"],
    "DATA ENTREGUE": ["dt_entrega"],
    "DATA CANCELAMENTO": ["dt_cancelamento"],
    "MOTIVO CANCELAMENTO": ["motivo_cancelamento"]
}

def formatar_excel(writer, sheet_name):
    ws = writer.sheets[sheet_name]
    for col_num, column_cells in enumerate(ws.iter_cols(1, ws.max_column), 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

def buscar_chave(d, chaves):
    for chave in chaves:
        if chave in d and d[chave]:
            return d[chave]
    return ""

def buscar_geo(sol, briefing):
    return (
        briefing.get("geo") or
        sol.get("geo") or
        briefing.get("unb") or
        sol.get("unb") or
        ""
    )

def buscar_unb(sol, briefing):
    return briefing.get("unb") or sol.get("unb") or ""

def buscar_sala(sol, briefing):
    return buscar_chave(briefing, ["sala"]) or buscar_chave(sol, ["sala"])

def buscar_setor(sol, briefing):
    return buscar_chave(briefing, ["setor"]) or buscar_chave(sol, ["setor"])

with st.sidebar:
    st.header("Credenciais e Configuração")
    st.session_state.api_key = "47200d78-1f10-44ba-a8e4-187a7d35e3bd"
    st.code("API Key usada:\n47200d78-1f10-44ba-a8e4-187a7d35e3bd", language="text")

opcao = st.radio("Selecione a operação:", ["Consultar Relatórios", "Consultar Detalhes da Solicitação"])

if opcao == "Consultar Relatórios":
    st.subheader("Parâmetros da Consulta")
    col1, col2 = st.columns(2)
    with col1:
        initial_date = st.date_input("Data Inicial")
    with col2:
        end_date = st.date_input("Data Final")
    briefing_id = st.text_input("Briefing ID (opcional)")
    board_name = st.text_input("Board (opcional)")
    column_name = st.text_input("Coluna (opcional)")

    if st.button("Consultar Relatórios"):
        with st.spinner("Consultando API da Deskfy..."):
            if not st.session_state.api_key or not initial_date or not end_date:
                st.warning("Preencha todos os campos obrigatórios.")
            else:
                url = "https://service-api.deskfy.io/v1/reports/workflow"
                params = {
                    "initialDate": initial_date.strftime("%Y-%m-%d"),
                    "endDate": end_date.strftime("%Y-%m-%d"),
                    "briefingId": briefing_id
                }
                if board_name:
                    params["boardName"] = board_name
                if column_name:
                    params["columnName"] = column_name

                response = requests.get(url, headers={"x-api-key": st.session_state.api_key}, params=params)
                if response.status_code == 200:
                    data = response.json()
                    if not data:
                        st.info("Nenhum dado encontrado.")
                    else:
                        df = pd.json_normalize(data)
                        if 'tags' in df.columns:
                            df['tags'] = df['tags'].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)
                        st.success("Dados obtidos com sucesso.")

                        st.markdown("### Resultado Formatado da Solicitação")
                        st.dataframe(df)

                        st.markdown("Copiar dados da tabela:")
                        st.code(df.to_csv(index=False, sep=';'), language='csv')

                        output_excel = BytesIO()
                        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name="Relatório")
                            formatar_excel(writer, "Relatório")
                        output_excel.seek(0)

                        st.download_button("Baixar Excel", data=output_excel, file_name="relatorio_deskfy.xlsx")
                        output_csv = df.to_csv(index=False).encode("utf-8")
                        st.download_button("Baixar CSV", data=output_csv, file_name="relatorio_deskfy.csv")
                else:
                    st.error(f"Erro {response.status_code}: {response.text}")

elif opcao == "Consultar Detalhes da Solicitação":
    st.subheader("Detalhes da Solicitação")
    task_id = st.text_input("Task ID da Solicitação")

    if st.button("Consultar Detalhes"):
        with st.spinner("Buscando dados da solicitação..."):
            if not (st.session_state.api_key and task_id):
                st.warning("Preencha a API Key e o Task ID.")
            else:
                url = "https://service-api.deskfy.io/v1/reports/workflow/task-details"
                response = requests.get(url, headers={"x-api-key": st.session_state.api_key}, params={"taskId": task_id})
                if response.status_code == 200:
                    data = response.json()
                    st.success("Dados carregados com sucesso.")

                    with st.expander("Ver JSON completo da resposta", expanded=False):
                        st.code(data, language="json")

                    sol = data.get("solicitacao", {})
                    briefing = data.get("briefing", {})
                    responsaveis = data.get("responsaveis", "")

                    if isinstance(responsaveis, list):
                        responsaveis = " | ".join(str(r) for r in responsaveis)

                    data_sol = buscar_chave(sol, ["dt_cadastro"])
                    data_entregue = buscar_chave(sol, ["dt_entrega"])

                    linha = {}
                    for col in COLUNAS_MODELO:
                        if col in CAMPOS_EXCEL_FORMULA:
                            linha[col] = ""
                        elif col == "GEO":
                            linha[col] = buscar_geo(sol, briefing)
                        elif col == "UNB":
                            linha[col] = buscar_unb(sol, briefing)
                        elif col == "SALA":
                            linha[col] = buscar_sala(sol, briefing)
                        elif col == "SETOR":
                            linha[col] = buscar_setor(sol, briefing)
                        elif col == "COMERCIAL":
                            linha[col] = responsaveis
                        elif col in mapeamento_briefing:
                            linha[col] = buscar_chave(briefing, mapeamento_briefing[col])
                        elif col in mapeamento_sol:
                            valor = buscar_chave(sol, mapeamento_sol[col])
                            if col == "LINK DESKFY":
                                linha[col] = f"https://app.deskfy.io/task/{valor}" if valor else ""
                            else:
                                linha[col] = valor
                        elif col == "DATA SOLICITAÇÃO":
                            linha[col] = data_sol
                        elif col == "MÊS SOLICITAÇÃO":
                            linha[col] = pd.to_datetime(data_sol).strftime("%m") if data_sol else ""
                        elif col == "ANO SOLICITAÇÃO":
                            linha[col] = pd.to_datetime(data_sol).strftime("%Y") if data_sol else ""
                        elif col == "DATA ENTREGUE":
                            linha[col] = data_entregue
                        elif col == "MÊS ENTREGUE":
                            linha[col] = pd.to_datetime(data_entregue).strftime("%m") if data_entregue else ""
                        elif col == "ANO ENTREGUE":
                            linha[col] = pd.to_datetime(data_entregue).strftime("%Y") if data_entregue else ""
                        else:
                            linha[col] = sol.get(col.lower(), briefing.get(col.lower(), data.get(col.lower(), "")))

                    df = pd.DataFrame([linha])
                    df = df[COLUNAS_MODELO]

                    st.markdown("### Resultado Formatado da Solicitação")
                    st.dataframe(df)

                    st.markdown("Copiar dados da tabela:")
                    st.code(df.to_csv(index=False, sep=';'), language='csv')

                    output = BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Detalhes")
                        formatar_excel(writer, "Detalhes")
                    output.seek(0)

                    st.download_button("Baixar Excel da Solicitação", data=output, file_name=f"detalhes_{linha['COLUNA DESKFY'] or task_id}.xlsx")
                else:
                    st.error(f"Erro Deskfy {response.status_code}: {response.text}")
                    